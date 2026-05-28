import os
import re
import json
import csv
from pathlib import Path
from typing import Tuple, List, Dict, Optional

import requests
from bs4 import BeautifulSoup
import PyPDF2
from docx import Document
from openpyxl import load_workbook
import xlrd

# -------------------------------------------------------------------
# Configuration
# -------------------------------------------------------------------
BASE_DIR = Path(r"C:\Scraping\Mississippi-Procurement\scraped_data")
OUTPUT_REPORT = Path("rfp_scan_report.json")
SCRAPED_LINKS_DIR = BASE_DIR / "_scraped_links"  # store scraped link data
REQUEST_TIMEOUT = 30
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"

# Keywords that indicate an RFP document (case‑insensitive)
RFP_KEYWORDS = [
    r"request\s+for\s+proposal",
    r"\brfp\b",
    r"request\s+for\s+quotation",
    r"\brfq\b",
    r"invitation\s+for\s+bids",
    r"\bifb\b",
    r"solicitation",
    r"procurement\s+opportunity",
    r"bid\s+opportunity",
]

HEADERS = {"User-Agent": USER_AGENT}

# -------------------------------------------------------------------
# Text extraction from various file types
# -------------------------------------------------------------------
def extract_text_from_text_file(filepath: Path) -> str:
    """Read plain text files (UTF-8, fallback to latin-1)."""
    encodings = ["utf-8", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            return filepath.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return ""

def extract_text_from_pdf(filepath: Path) -> str:
    """Extract text from PDF using PyPDF2."""
    text = []
    try:
        with open(filepath, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text.append(page_text)
    except Exception as e:
        print(f"  [ERROR] PDF extraction failed {filepath.name}: {e}")
    return "\n".join(text)

def extract_text_from_docx(filepath: Path) -> str:
    """Extract text from .docx file."""
    text = []
    try:
        doc = Document(filepath)
        for para in doc.paragraphs:
            text.append(para.text)
    except Exception as e:
        print(f"  [ERROR] DOCX extraction failed {filepath.name}: {e}")
    return "\n".join(text)

def extract_text_from_xlsx(filepath: Path) -> str:
    """Extract text from .xlsx using openpyxl."""
    text = []
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text.append(str(cell))
    except Exception as e:
        print(f"  [ERROR] XLSX extraction failed {filepath.name}: {e}")
    return "\n".join(text)

def extract_text_from_xls(filepath: Path) -> str:
    """Extract text from old .xls using xlrd."""
    text = []
    try:
        wb = xlrd.open_workbook(filepath, ignore_workbook_corruption=True)
        for sheet in wb.sheets():
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    val = sheet.cell_value(row, col)
                    if val:
                        text.append(str(val))
    except Exception as e:
        print(f"  [ERROR] XLS extraction failed {filepath.name}: {e}")
    return "\n".join(text)

def extract_text_from_csv(filepath: Path) -> str:
    """Extract text from CSV."""
    text = []
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for row in reader:
                text.extend([str(cell) for cell in row if cell])
    except Exception as e:
        print(f"  [ERROR] CSV extraction failed {filepath.name}: {e}")
    return "\n".join(text)

def extract_text_from_file(filepath: Path) -> str:
    """Route file to appropriate extractor based on extension."""
    ext = filepath.suffix.lower()
    if ext in [".txt", ".json", ".html", ".htm", ".xml", ".md"]:
        return extract_text_from_text_file(filepath)
    elif ext == ".pdf":
        return extract_text_from_pdf(filepath)
    elif ext == ".docx":
        return extract_text_from_docx(filepath)
    elif ext == ".xlsx":
        return extract_text_from_xlsx(filepath)
    elif ext == ".xls":
        return extract_text_from_xls(filepath)
    elif ext == ".csv":
        return extract_text_from_csv(filepath)
    else:
        # For unknown binary files, just return empty string
        print(f"  [SKIP] Unsupported file type: {ext}")
        return ""

# -------------------------------------------------------------------
# RFP detection
# -------------------------------------------------------------------
def is_rfp_content(text: str) -> bool:
    """Return True if any RFP keyword matches the text (case‑insensitive)."""
    if not text:
        return False
    text_lower = text.lower()
    for pattern in RFP_KEYWORDS:
        if re.search(pattern, text_lower):
            return True
    return False

# -------------------------------------------------------------------
# URL extraction from text
# -------------------------------------------------------------------
def extract_urls(text: str) -> List[str]:
    """Find all http/https URLs in the text."""
    url_pattern = r"https?://(?:[-\w.]|(?:%[\da-fA-F]{2}))+(?:/[-\w%!$&'()*+,;=:@/]*)*"
    urls = re.findall(url_pattern, text)
    # Remove duplicates while preserving order
    seen = set()
    unique_urls = []
    for url in urls:
        if url not in seen:
            seen.add(url)
            unique_urls.append(url)
    return unique_urls

# -------------------------------------------------------------------
# Scrape a single URL (full page content)
# -------------------------------------------------------------------
def scrape_url(url: str) -> Optional[str]:
    """Download and extract all visible text from a URL."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        # Remove script, style, meta, etc.
        for tag in soup(["script", "style", "meta", "noscript"]):
            tag.decompose()
        text = soup.get_text(separator="\n", strip=True)
        return text
    except Exception as e:
        print(f"    [SCRAPE ERROR] {url[:80]}: {e}")
        return None

# -------------------------------------------------------------------
# Main scanning routine
# -------------------------------------------------------------------
def scan_directory(base_dir: Path):
    """Walk through all subdirectories, process each file."""
    if not base_dir.exists():
        print(f"ERROR: Directory not found: {base_dir}")
        return

    # Prepare output directory for scraped link data
    SCRAPED_LINKS_DIR.mkdir(parents=True, exist_ok=True)

    report = {
        "scan_root": str(base_dir.resolve()),
        "files_processed": 0,
        "rfp_files": [],
        "non_rfp_files": [],
        "links_found": [],
        "scraped_links": {},
    }

    # Walk through all subdirectories (including base_dir itself)
    for root, dirs, files in os.walk(base_dir):
        root_path = Path(root)
        # Skip the _scraped_links folder we create
        if root_path == SCRAPED_LINKS_DIR:
            continue
        for filename in files:
            filepath = root_path / filename
            rel_path = filepath.relative_to(base_dir)
            print(f"\n📄 Processing: {rel_path}")

            # 1. Extract text content
            text = extract_text_from_file(filepath)
            if not text.strip():
                print(f"   -> No extractable text (or file is binary/empty)")
                # Still record it but mark as non‑rfp
                report["non_rfp_files"].append(str(rel_path))
                report["files_processed"] += 1
                continue

            # 2. Detect RFP
            rfp_flag = is_rfp_content(text)
            classification = "RFP" if rfp_flag else "Non‑RFP"
            print(f"   -> Classification: {classification}")

            # 3. Find links inside the file
            urls = extract_urls(text)
            if urls:
                print(f"   -> Found {len(urls)} URL(s):")
                for u in urls:
                    print(f"       {u[:100]}")
                report["links_found"].append({
                    "file": str(rel_path),
                    "urls": urls
                })
            else:
                print(f"   -> No URLs found.")

            # 4. Scrape each link (only if we haven't scraped it before)
            scraped_data_for_file = {}
            for url in urls:
                # Use a unique key: URL (could also add date, but simple dedup)
                if url not in report["scraped_links"]:
                    print(f"   -> Scraping: {url[:80]}...")
                    scraped_text = scrape_url(url)
                    if scraped_text:
                        report["scraped_links"][url] = {
                            "url": url,
                            "source_file": str(rel_path),
                            "content_preview": scraped_text[:500],
                            "full_text_length": len(scraped_text)
                        }
                        # Also save the full scraped text to a file for later inspection
                        safe_name = re.sub(r"[^\w\-]", "_", url)[:100]
                        out_file = SCRAPED_LINKS_DIR / f"{safe_name}.txt"
                        out_file.write_text(scraped_text, encoding="utf-8")
                        print(f"       ✓ Saved full scraped content to {out_file.name}")
                    else:
                        report["scraped_links"][url] = {
                            "url": url,
                            "source_file": str(rel_path),
                            "error": "Failed to scrape"
                        }
                else:
                    print(f"   -> Already scraped: {url[:80]} (skip)")

            # 5. Update report with this file's classification
            file_info = {
                "path": str(rel_path),
                "classification": classification,
                "urls_found": urls,
                "text_preview": text[:300]
            }
            if rfp_flag:
                report["rfp_files"].append(file_info)
            else:
                report["non_rfp_files"].append(file_info)

            report["files_processed"] += 1

    # Write final report
    OUTPUT_REPORT.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    print("\n" + "="*60)
    print(f"✅ SCAN COMPLETE")
    print(f"   Files processed    : {report['files_processed']}")
    print(f"   RFP files          : {len(report['rfp_files'])}")
    print(f"   Non‑RFP files      : {len(report['non_rfp_files'])}")
    print(f"   Unique links found : {len(report['links_found'])} (total across files)")
    print(f"   Links successfully scraped: {len([v for v in report['scraped_links'].values() if 'error' not in v])}")
    print(f"   Report saved to    : {OUTPUT_REPORT.resolve()}")
    print(f"   Scraped link data  : {SCRAPED_LINKS_DIR.resolve()}")
    print("="*60)

# -------------------------------------------------------------------
# Entry point
# -------------------------------------------------------------------
if __name__ == "__main__":
    scan_directory(BASE_DIR)